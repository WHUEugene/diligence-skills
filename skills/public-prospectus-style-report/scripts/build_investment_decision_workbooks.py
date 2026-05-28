#!/usr/bin/env python3
"""Export investment-decision diligence matrices from JSON to XLSX workbooks."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SHEETS = {
    "claims": {
        "filename": "01_PPT主张清单.xlsx",
        "title": "PPT主张清单",
        "fields": [
            "slide_no",
            "ppt_section",
            "original_claim",
            "claim_type",
            "quantitative_value",
            "report_section",
            "evidence_required",
            "external_verification_needed",
            "risk_level",
        ],
        "headers": [
            "页码/位置",
            "PPT章节",
            "项目方原始主张",
            "主张类型",
            "量化数值",
            "报告落点",
            "所需证据",
            "是否需外部验证",
            "风险等级",
        ],
    },
    "diligence_requests": {
        "filename": "02_尽调问题与资料索取清单.xlsx",
        "title": "尽调问题与资料索取清单",
        "fields": [
            "question_id",
            "report_section",
            "diligence_question",
            "required_document",
            "owner_or_source",
            "why_it_matters",
            "decision_use",
            "priority",
        ],
        "headers": [
            "编号",
            "报告章节",
            "尽调问题",
            "需取得资料",
            "责任方/来源",
            "重要性",
            "投决用途",
            "优先级",
        ],
    },
    "evidence_matrix": {
        "filename": "03_证据矩阵.xlsx",
        "title": "证据矩阵",
        "fields": [
            "ppt_claim",
            "company_evidence",
            "interview_evidence",
            "external_evidence",
            "evidence_strength",
            "conclusion",
            "report_sentence",
            "follow_up_question",
        ],
        "headers": [
            "项目方主张",
            "企业资料证据",
            "访谈证据",
            "外部证据",
            "证据强度",
            "报告结论",
            "可入正文句子",
            "后续核验问题",
        ],
    },
    "risks_conditions": {
        "filename": "04_风险与前置条件清单.xlsx",
        "title": "风险与前置条件清单",
        "fields": [
            "risk_item",
            "fact_basis",
            "risk_judgment",
            "investment_impact",
            "condition_precedent",
            "control_clause",
            "verification_document",
            "report_section",
        ],
        "headers": [
            "风险事项",
            "事实基础",
            "风险判断",
            "投资影响",
            "前置条件",
            "风控条款",
            "核验资料",
            "报告章节",
        ],
    },
}


def normalize_rows(value) -> list[dict]:
    if value is None:
        return []
    if not isinstance(value, list):
        raise TypeError("matrix sections must be lists of objects")
    rows = []
    for idx, item in enumerate(value, 1):
        if not isinstance(item, dict):
            raise TypeError(f"row {idx} is not an object")
        rows.append(item)
    return rows


def value_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return "\n".join(value_to_text(v) for v in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def style_sheet(ws, headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F3A63")
    header_font = Font(name="宋体", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="宋体", size=10, color="000000")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    for idx, _header in enumerate(headers, 1):
        letter = get_column_letter(idx)
        max_len = 12
        for cell in ws[letter]:
            text = value_to_text(cell.value)
            max_len = max(max_len, min(48, max((len(line) for line in text.splitlines()), default=0) + 2))
        ws.column_dimensions[letter].width = max_len
    ws.row_dimensions[1].height = 28


def write_workbook(output_path: Path, title: str, fields: list[str], headers: list[str], rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(headers)
    for row in rows:
        ws.append([value_to_text(row.get(field, "")) for field in fields])
    style_sheet(ws, headers)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input_json", help="JSON file containing claims, diligence_requests, evidence_matrix, risks_conditions")
    parser.add_argument("--output-dir", help="Directory for XLSX outputs. Defaults to the JSON file directory.")
    parser.add_argument("--strict", action="store_true", help="Fail if any required matrix is empty.")
    args = parser.parse_args()

    input_path = Path(args.input_json)
    data = json.loads(input_path.read_text(encoding="utf-8"))
    output_dir = Path(args.output_dir) if args.output_dir else input_path.parent

    manifest = {}
    for key, spec in SHEETS.items():
        rows = normalize_rows(data.get(key))
        if args.strict and not rows:
            raise SystemExit(f"matrix `{key}` is empty")
        out = output_dir / spec["filename"]
        write_workbook(out, spec["title"], spec["fields"], spec["headers"], rows)
        manifest[key] = str(out)

    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
